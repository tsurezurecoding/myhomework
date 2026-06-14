# -*- coding: utf-8 -*-
import argparse
from datetime import date, datetime
import json
import re
import tempfile
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_SPREADSHEET_ID = "1IeBaoI0xaE_jQO9TXZBMiEWIoLdxb9tNSdEeCh9Rjbc"
DEFAULT_OUT = Path(__file__).resolve().parent / "app" / "src" / "features" / "app-data.js"
NOTE_HEADER_RE = re.compile(r"備考|メモ")
DATE_HEADER_RE = re.compile(r"^日付$|予定")
COMPLETED_VALUES = {"〇", "○", "完了", "見直し完了"}
DATED_COMPLETED_RE = re.compile(r"^\d{4}[/-]\d{1,2}[/-]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?\s*[〇○]$")


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def is_completed_value(value):
    text = cell_text(value)
    return text in COMPLETED_VALUES or bool(DATED_COMPLETED_RE.match(text))


def planned_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text or is_completed_value(value):
        return ""

    normalized = text.replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m-%d"):
        try:
            parsed = datetime.strptime(normalized, fmt)
            if fmt == "%m-%d":
                parsed = parsed.replace(year=date.today().year)
            return parsed.date().isoformat()
        except ValueError:
            continue
    return ""


def completed_date(value):
    """「日付＋〇」で完了したセルから完了日(ISO)を返す。日付なし完了（〇・見直し完了等）は空。"""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = cell_text(value)
    if not DATED_COMPLETED_RE.match(text):
        return ""
    match = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})", text)
    if not match:
        return ""
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
    except ValueError:
        return ""


def header_column(sheet, pattern):
    """ヘッダー行(1行目)から pattern に一致する列番号を返す（無ければ None）。"""
    for column in range(1, sheet.max_column + 1):
        if pattern.search(cell_text(sheet.cell(1, column).value)):
            return column
    return None


def download_sheet(spreadsheet_id):
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=30) as response:
        data = response.read()
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp.write(data)
    temp.close()
    return Path(temp.name)


def parse_args():
    parser = argparse.ArgumentParser(description="Generate static app data from the homework Google Sheet.")
    parser.add_argument("--spreadsheet-id", default=DEFAULT_SPREADSHEET_ID)
    parser.add_argument("--xlsx", type=Path, default=None)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    return parser.parse_args()


def main():
    args = parse_args()
    source = args.xlsx or download_sheet(args.spreadsheet_id)
    wb = load_workbook(source, data_only=True)
    items = []
    subject_counts = {}

    for sheet in wb.worksheets:
        subject = sheet.title
        subject_counts[subject] = 0
        note_col = header_column(sheet, NOTE_HEADER_RE)
        date_col = header_column(sheet, DATE_HEADER_RE)
        for row_number in range(2, sheet.max_row + 1):
            values = [sheet.cell(row_number, column).value for column in range(1, 8)]
            if not any(values):
                continue
            lesson_progress = cell_text(values[0])
            raw_status = cell_text(values[6])
            status = "completed" if is_completed_value(values[6]) else "remaining"
            due_date = planned_date(sheet.cell(row_number, date_col).value) if date_col else planned_date(values[6])
            done_date = completed_date(values[6]) if status == "completed" else ""
            note = cell_text(sheet.cell(row_number, note_col).value) if note_col else ""
            subject_counts[subject] += 1
            items.append(
                {
                    "id": f"{subject}-{subject_counts[subject]}",
                    "subject": subject,
                    "lessonProgress": "done" if is_completed_value(values[0]) else "notYet",
                    "lessonProgressRaw": lesson_progress,
                    "textbook": values[1] or "",
                    "textbookRange": values[2] or "",
                    "testRange": values[3] or "",
                    "material": values[4] or "",
                    "task": values[5] or "",
                    "status": status,
                    "rawStatus": raw_status,
                    "plannedDate": due_date,
                    "completedDate": done_date,
                    "note": note,
                }
            )

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w", encoding="utf-8", newline="\n") as file:
        file.write("window.HOMEWORK_ITEMS = ")
        json.dump(items, file, ensure_ascii=False, separators=(",", ":"))
        file.write(";\n")
    print(f"{len(items)} items written to {args.out}")


if __name__ == "__main__":
    main()
